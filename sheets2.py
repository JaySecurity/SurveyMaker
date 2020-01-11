import csv
import openpyxl
from openpyxl.styles import Border, Side, Alignment, NamedStyle
from openpyxl.worksheet.pagebreak import Break

infile = open('sheets.csv', 'r')
COLWIDTH = {'A':6.5,'B':15,'C':13,'D':35,'E':6,'F':6,'G':6,'H':8.5,'I':8.5,'J':8.5,'K':8.5, 'L':11}
ALIGNMENT = Alignment(horizontal='general', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
datafield = NamedStyle(name='datafield')
hourfield = NamedStyle(name='hourfield')
reader = csv.reader(infile)
#Create Wookbook

wb = openpyxl.Workbook()
ws = wb.active
bd = Side(style='thin', color="000000")
datafield.border = Border(left=Side(border_style='thin', color = '000000'), right=Side(border_style='thin', color = '000000'), top=Side(border_style='thin', color = '000000'), bottom=Side(border_style='thin', color = '000000'))
hourfield.border = Border(bottom= bd)
    
wb.add_named_style(datafield)
wb.add_named_style(hourfield)
#Setup WorkSheet

ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_margins.left = .25
ws.page_margins.right = .25
ws.page_margins.top = .5
ws.page_margins.bottom = .5
wsprops = ws.sheet_properties

#Set Column Widths

for k,v in COLWIDTH.items():
    ws.column_dimensions[k].width = v

i=0
unit = 0
for row in reader:
    newrow = []
    
    if row[0].startswith('CA836'):
        if unit == 3:
            ws.row_breaks.append(Break(id=i))
            unit = 0
        i+=1
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=10)
        ws.cell(i,11,"New Hours")
        ws.cell(i,1, value = row[0])
        ws.cell(i, 12).style = 'hourfield'
        i += 1
        ws.append(['Pos', 'Serial #', 'Brand', 'Description', 'Prev PSI', 'New PSI', 'Hot / Cold', 'Prev Outer TD', 'Current Outer TD', 'Prev Inner TD', 'Current Inner TD'])
        ws.row_dimensions[i].height = 30
        for c in range(5,12):
            ws.cell(row=i,column=c).alignment = ALIGNMENT
        i += 1
        unit +=1
    if row[0].startswith('P'):
        position = row[0].split('-')
        ws.append([position[-1], row[1], row[2], row[3], row[5], '', '',row[8],'', row[9],''])
        for c in range(5,12):
            ws.cell(row=i,column=c).style = 'datafield'
        i +=1

infile.close
wb.save('survey.xlsx')
